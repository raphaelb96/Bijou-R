from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Palette ───────────────────────────────────────────────────────────────────
OR          = "FFD700"
OR_LIGHT    = "FFF8DC"
OR_MED      = "FFE680"
NOIR        = "1A1A1A"
GRIS        = "F2F2F2"
GRIS_MED    = "D9D9D9"
BLANC       = "FFFFFF"
VERT        = "2E7D32"
VERT_LIGHT  = "E8F5E9"
BLEU        = "1565C0"
BLEU_LIGHT  = "E3F2FD"
ROUGE       = "B71C1C"
ROUGE_LIGHT = "FFEBEE"
VIOLET      = "6A1B9A"
VIOLET_LIGHT= "F3E5F5"
TEAL        = "00695C"
TEAL_LIGHT  = "E0F2F1"

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color=NOIR, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def thin(): s=Side(style="thin",color=GRIS_MED); return Border(left=s,right=s,top=s,bottom=s)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def hdr(ws, r, c, v, bg=NOIR, fg=BLANC, sz=10, bold=True):
    x = ws.cell(row=r, column=c, value=v)
    x.fill=fill(bg); x.font=fnt(bold,fg,sz); x.alignment=ctr(); x.border=thin()
    return x

def cel(ws, r, c, v=None, bg=BLANC, fg=NOIR, bold=False, align="left",
        italic=False, sz=10, fmt=None):
    x = ws.cell(row=r, column=c, value=v)
    x.fill=fill(bg); x.font=fnt(bold,fg,sz,italic)
    x.alignment=ctr() if align=="center" else lft()
    x.border=thin()
    if fmt: x.number_format=fmt
    return x

# ─────────────────────────────────────────────────────────────────────────────
# TABLE POIDS : consolidation des 2 fichiers de référence
# Source 1 : TAMISSIZEWEIGHT (standard industrie)
# Source 2 : DIAMOND_sizeweight (incréments fins 0.05mm)
# ─────────────────────────────────────────────────────────────────────────────
TABLE_POIDS = [
    # mm,    ct
    (0.80,  0.0025),
    (0.85,  0.0023),   # source 2
    (0.90,  0.005),
    (0.95,  0.0032),   # source 2
    (1.00,  0.006),
    (1.05,  0.00429),  # source 2
    (1.10,  0.00654),
    (1.15,  0.00712),
    (1.20,  0.00815),
    (1.25,  0.009305),
    (1.30,  0.01010),
    (1.35,  0.01158),
    (1.40,  0.01225),
    (1.45,  0.01429),
    (1.50,  0.01522),
    (1.55,  0.01721),
    (1.60,  0.01875),
    (1.65,  0.01689),  # source 2
    (1.70,  0.02318),
    (1.75,  0.02011),  # source 2
    (1.80,  0.02742),
    (1.85,  0.02378),  # source 2
    (1.90,  0.02940),
    (1.95,  0.02777),  # source 2
    (2.00,  0.03624),
    (2.10,  0.04500),
    (2.20,  0.05000),
    (2.30,  0.05500),
    (2.40,  0.06000),
    (2.50,  0.06600),
    (2.60,  0.07400),
    (2.70,  0.08000),
    (2.80,  0.09000),
    (2.90,  0.10000),
    (3.00,  0.11000),
    (3.10,  0.12000),
    (3.20,  0.13500),
    (3.30,  0.15000),
    (3.40,  0.16000),
    (3.50,  0.17000),
    (3.60,  0.18000),
    (3.70,  0.20000),
    (3.80,  0.22000),
    (3.90,  0.24000),
    (4.00,  0.25000),
    (4.10,  0.26000),
    (4.20,  0.28000),
    (4.30,  0.30000),
    (4.40,  0.33000),
    (4.50,  0.35000),
    (4.75,  0.44000),
    (5.00,  0.46000),
    (5.20,  0.52000),  # interpolé
    (5.25,  0.56000),
    (5.50,  0.66000),
    (5.75,  0.76000),
    (6.00,  0.84000),
    (6.25,  0.93000),
    (6.50,  1.03000),
    (6.75,  1.24000),
    (7.00,  1.28000),
    (7.25,  1.49000),
    (7.50,  1.67000),
    (7.75,  1.75000),
    (8.00,  2.04000),
    (8.25,  2.11000),
    (8.50,  2.43000),
    (8.75,  2.55000),
    (9.00,  2.75000),
    (9.25,  3.05000),
    (9.50,  3.35000),
    (9.75,  3.61000),
    (10.00, 3.87000),
    (10.25, 4.16000),
    (10.50, 4.41000),
    (10.75, 4.57000),
    (11.00, 4.91000),
    (11.25, 5.49000),
    (11.50, 5.85000),
    (11.75, 6.47000),
    (12.00, 6.84000),
    (12.25, 7.26000),
    (12.50, 7.36000),
    (12.75, 7.52000),
    (13.00, 8.51000),
    (13.50, 9.53000),
    (14.00, 10.49000),
    (15.00, 12.89000),
    (16.00, 16.16000),
]
TABLE_POIDS.sort(key=lambda x: x[0])

# ═══════════════════════════════════════════════════════════════════════════════
# ONGLET 1 — TABLE_POIDS
# ═══════════════════════════════════════════════════════════════════════════════
ws_tp = wb.active
ws_tp.title = "TABLE_POIDS"
ws_tp.sheet_properties.tabColor = TEAL

ws_tp.column_dimensions["A"].width = 18
ws_tp.column_dimensions["B"].width = 18
ws_tp.column_dimensions["C"].width = 40

ws_tp.merge_cells("A1:C1")
c = ws_tp["A1"]
c.value = "TABLE DE RÉFÉRENCE — Diamètre (mm) → Poids (carats)"
c.fill = fill(TEAL); c.font = fnt(True, BLANC, 13); c.alignment = ctr()
ws_tp.row_dimensions[1].height = 32

ws_tp.merge_cells("A2:C2")
c = ws_tp["A2"]
c.value = ("Sources : TAMISSIZEWEIGHT (standard industrie) + DIAMOND_sizeweight. "
           "Ne pas modifier — utilisée par VLOOKUP dans le CATALOGUE.")
c.fill = fill(TEAL_LIGHT); c.font = fnt(False, NOIR, 9, True); c.alignment = lft()
ws_tp.row_dimensions[2].height = 20

hdr(ws_tp, 3, 1, "Diamètre (mm)", bg=TEAL, fg=BLANC)
hdr(ws_tp, 3, 2, "Poids unitaire (ct)", bg=TEAL, fg=BLANC)
hdr(ws_tp, 3, 3, "Indication visuelle", bg=TEAL, fg=BLANC)
ws_tp.row_dimensions[3].height = 20

indications = {
    (0.00, 0.90): "Melee très fin",
    (0.90, 1.50): "Melee standard",
    (1.50, 3.00): "Petit diamant",
    (3.00, 5.00): "Diamant moyen",
    (5.00, 8.00): "Beau diamant",
    (8.00, 99.0): "Grand diamant / Pierre centrale",
}
def get_indication(mm):
    for (lo, hi), label in indications.items():
        if lo < mm <= hi or (lo == 0.00 and mm <= hi):
            return label
    return ""

for i, (mm, ct) in enumerate(TABLE_POIDS, 4):
    bg = TEAL_LIGHT if i % 2 == 0 else BLANC
    c = cel(ws_tp, i, 1, mm, bg=bg, align="center", fmt='0.00 "mm"')
    c.font = fnt(True, TEAL, 10)
    c = cel(ws_tp, i, 2, ct, bg=bg, align="center", fmt='0.00000 "ct"')
    cel(ws_tp, i, 3, get_indication(mm), bg=bg, italic=True)
    ws_tp.row_dimensions[i].height = 16

# ═══════════════════════════════════════════════════════════════════════════════
# ONGLET 2 — COURS
# ═══════════════════════════════════════════════════════════════════════════════
ws_cours = wb.create_sheet("COURS")
ws_cours.sheet_properties.tabColor = OR

ws_cours.column_dimensions["A"].width = 35
ws_cours.column_dimensions["B"].width = 18
ws_cours.column_dimensions["C"].width = 14
ws_cours.column_dimensions["D"].width = 42

ws_cours.merge_cells("A1:D1")
c = ws_cours["A1"]
c.value = "COURS DU MARCHÉ — BIJOU-R"
c.fill = fill(NOIR); c.font = fnt(True, OR, 14); c.alignment = ctr()
ws_cours.row_dimensions[1].height = 35

ws_cours.merge_cells("A2:D2")
c = ws_cours["A2"]
c.value = "⚠  Mettez à jour les cellules jaunes pour recalculer automatiquement tous les prix"
c.fill = fill(OR_LIGHT); c.font = fnt(True, NOIR, 10); c.alignment = ctr()
ws_cours.row_dimensions[2].height = 22

sections_cours = [
    ("OR", OR, NOIR, [
        ("Cours de l'or", 85.00, "€ / gramme",
         "Prix spot international. Consulter kitco.com ou goldprice.org"),
    ]),
    ("DIAMANTS NATURELS", BLEU, BLANC, [
        ("Naturel — N1 Essentiel",   1800,  "€ / carat", "H-I / SI1-SI2"),
        ("Naturel — N2 Sélection",   3500,  "€ / carat", "G-H / VS2"),
        ("Naturel — N3 Prestige",    6200,  "€ / carat", "F-G / VS1"),
        ("Naturel — N4 Excellence",  10500, "€ / carat", "E-F / VVS1-VVS2"),
        ("Naturel — N5 Exception",   18000, "€ / carat", "D / IF — collector"),
    ]),
    ("DIAMANTS DE LABORATOIRE", VIOLET, BLANC, [
        ("Labo — L1 Essentiel",  420,  "€ / carat", "H-I / SI1-SI2"),
        ("Labo — L2 Sélection",  780,  "€ / carat", "G-H / VS2"),
        ("Labo — L3 Prestige",   1300, "€ / carat", "F-G / VS1"),
        ("Labo — L4 Excellence", 2100, "€ / carat", "E-F / VVS1-VVS2"),
        ("Labo — L5 Exception",  3500, "€ / carat", "D / IF"),
    ]),
]

row = 3
for title, bg, fg, params in sections_cours:
    ws_cours.merge_cells(f"A{row}:D{row}")
    c = ws_cours[f"A{row}"]
    c.value = title; c.fill = fill(bg); c.font = fnt(True, fg, 11); c.alignment = ctr()
    ws_cours.row_dimensions[row].height = 20; row += 1

    for col, label in enumerate(["Paramètre","Valeur","Unité","Note"], 1):
        hdr(ws_cours, row, col, label, bg=bg, fg=fg)
    ws_cours.row_dimensions[row].height = 18; row += 1

    for param, val, unit, note in params:
        value_bg = OR if bg == OR else (BLEU_LIGHT if bg == BLEU else VIOLET_LIGHT)
        cel(ws_cours, row, 1, param, bg=value_bg, bold=True)
        c = cel(ws_cours, row, 2, val, bg=OR if bg==OR else value_bg,
                bold=True, align="center", fmt='#,##0 "€"')
        cel(ws_cours, row, 3, unit, bg=value_bg, align="center")
        cel(ws_cours, row, 4, note, bg=value_bg, italic=True)
        ws_cours.row_dimensions[row].height = 18; row += 1

# ═══════════════════════════════════════════════════════════════════════════════
# ONGLET 3 — CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
ws_cfg = wb.create_sheet("CONFIG")
ws_cfg.sheet_properties.tabColor = GRIS_MED

ws_cfg.column_dimensions["A"].width = 35
ws_cfg.column_dimensions["B"].width = 18
ws_cfg.column_dimensions["C"].width = 14
ws_cfg.column_dimensions["D"].width = 42

ws_cfg.merge_cells("A1:D1")
c = ws_cfg["A1"]
c.value = "CONFIGURATION — PARAMÈTRES FIXES"
c.fill = fill(NOIR); c.font = fnt(True, BLANC, 14); c.alignment = ctr()
ws_cfg.row_dimensions[1].height = 35

sections_cfg = [
    ("OR & MÉTAUX", NOIR, BLANC, [
        ("Pureté or 18k",                  0.75, "facteur",   "75% or pur dans alliage 18k"),
        ("Supplément or blanc (rhodium)",   0.08, "facteur +%","Coût plaquage rhodium"),
        ("Supplément or rose",              0.02, "facteur +%","Alliage cuivre"),
    ]),
    ("FABRICATION", VERT, BLANC, [
        ("Main d'œuvre par défaut",  180.00, "€", "Modifiable ligne par ligne dans le catalogue"),
        ("Frais fixes (écrin, cert)", 25.00, "€", "Emballage, certificat, écrin"),
    ]),
    ("TARIFICATION", ROUGE, BLANC, [
        ("Coefficient de marge", 2.8,  "×",  "Prix HT = prix revient × coefficient"),
        ("TVA",                  0.20, "%",  "20% — France"),
    ]),
]

# Mapping pour les formules du CATALOGUE
# CONFIG!B5 = pureté or
# CONFIG!B6 = supplément or blanc
# CONFIG!B7 = supplément or rose
# CONFIG!B9 = main d'oeuvre défaut
# CONFIG!B10= frais fixes
# CONFIG!B12= coefficient marge
# CONFIG!B13= TVA

row = 2
for title, bg, fg, params in sections_cfg:
    ws_cfg.merge_cells(f"A{row}:D{row}")
    c = ws_cfg[f"A{row}"]
    c.value = title; c.fill = fill(bg); c.font = fnt(True, fg, 11); c.alignment = ctr()
    ws_cfg.row_dimensions[row].height = 20; row += 1

    for col, label in enumerate(["Paramètre","Valeur","Unité","Note"], 1):
        hdr(ws_cfg, row, col, label, bg=bg, fg=fg)
    ws_cfg.row_dimensions[row].height = 18; row += 1

    for param, val, unit, note in params:
        cel(ws_cfg, row, 1, param, bg=GRIS, bold=True)
        fmt = '#,##0.00 "€"' if unit == "€" else ("0%" if "%" in unit else "0.00")
        c = cel(ws_cfg, row, 2, val, bg=OR_LIGHT if bg==NOIR else GRIS,
                bold=True, align="center", fmt=fmt)
        cel(ws_cfg, row, 3, unit, bg=GRIS, align="center")
        cel(ws_cfg, row, 4, note, bg=GRIS, italic=True)
        ws_cfg.row_dimensions[row].height = 18; row += 1

# ═══════════════════════════════════════════════════════════════════════════════
# ONGLET 4 — CATALOGUE
# Colonnes :
#  A  REF       B  NOM         C  COLLECTION  D  MÉTAL
#  E  TYPE_DIAM F  NIVEAU_DIAM
#  G  G1_mm  H  G1_qté   I  G2_mm  J  G2_qté   K  G3_mm  L  G3_qté
#  M  G4_mm  N  G4_qté   O  G5_mm  P  G5_qté
#  Q  CARATS_CALC (auto)  R  POIDS_OR  S  MAIN_OEUVRE
#  T  COUT_OR   U  COUT_DIAM  V  REVIENT  W  PRIX_HT  X  PRIX_TTC
# ═══════════════════════════════════════════════════════════════════════════════
ws_cat = wb.create_sheet("CATALOGUE")
ws_cat.sheet_properties.tabColor = OR

COL = dict(
    REF=1, NOM=2, COLL=3, METAL=4, TYPE=5, NIV=6,
    G1MM=7,  G1QT=8,
    G2MM=9,  G2QT=10,
    G3MM=11, G3QT=12,
    G4MM=13, G4QT=14,
    G5MM=15, G5QT=16,
    CARATS=17, POIDS_OR=18, MO=19,
    COUT_OR=20, COUT_DIAM=21, REVIENT=22, HT=23, TTC=24,
)

col_widths = [14,26,16,12,12,18, 9,8, 9,8, 9,8, 9,8, 9,8, 13,12,14, 13,13,14,13,14]
for i, w in enumerate(col_widths, 1):
    ws_cat.column_dimensions[get_column_letter(i)].width = w

# Titre
ws_cat.merge_cells(f"A1:{get_column_letter(24)}1")
c = ws_cat["A1"]
c.value = "CATALOGUE PRODUITS — BIJOU-R"
c.fill = fill(NOIR); c.font = fnt(True, OR, 14); c.alignment = ctr()
ws_cat.row_dimensions[1].height = 35

# Groupes d'en-têtes ligne 2
grp2 = [
    (1,  3,  "IDENTITÉ PRODUIT",                    NOIR,  BLANC),
    (4,  6,  "MATIÈRES",                             BLEU,  BLANC),
    (7,  16, "GROUPES DIAMANTS  (mm × quantité)",    TEAL,  BLANC),
    (17, 19, "POIDS & FABRICATION",                  VERT,  BLANC),
    (20, 21, "COÛTS MATIÈRES",                       ROUGE, BLANC),
    (22, 24, "PRIX CALCULÉS  ✦  ne pas modifier",    OR,    NOIR),
]
for start, end, label, bg, fg in grp2:
    if start != end:
        ws_cat.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
    c = ws_cat.cell(row=2, column=start, value=label)
    c.fill=fill(bg); c.font=fnt(True,fg,9); c.alignment=ctr(); c.border=thin()
ws_cat.row_dimensions[2].height = 22

# En-têtes colonnes ligne 3
hdrs3 = [
    ("Référence",     NOIR,  BLANC),
    ("Nom modèle",    NOIR,  BLANC),
    ("Collection",    NOIR,  BLANC),
    ("Métal",         BLEU,  BLANC),
    ("Type diamant",  BLEU,  BLANC),
    ("Niveau",        BLEU,  BLANC),
    ("G1 mm",   TEAL, BLANC), ("G1 qté", TEAL, BLANC),
    ("G2 mm",   TEAL, BLANC), ("G2 qté", TEAL, BLANC),
    ("G3 mm",   TEAL, BLANC), ("G3 qté", TEAL, BLANC),
    ("G4 mm",   TEAL, BLANC), ("G4 qté", TEAL, BLANC),
    ("G5 mm",   TEAL, BLANC), ("G5 qté", TEAL, BLANC),
    ("Carats\n(calculés)", VERT, BLANC),
    ("Poids\nor (g)",      VERT, BLANC),
    ("Main\nd'œuvre €",    VERT, BLANC),
    ("Coût\nor €",   ROUGE, BLANC),
    ("Coût\ndiam €", ROUGE, BLANC),
    ("Revient\n€",   OR, NOIR),
    ("Prix\nHT €",   OR, NOIR),
    ("Prix\nTTC €",  OR, NOIR),
]
for i, (h, bg, fg) in enumerate(hdrs3, 1):
    hdr(ws_cat, 3, i, h, bg=bg, fg=fg, sz=8)
ws_cat.row_dimensions[3].height = 36
ws_cat.freeze_panes = "A4"

# ── Validations déroulantes ───────────────────────────────────────────────────
dv_metal = DataValidation(type="list",
    formula1='"Or blanc,Or jaune,Or rose"', showDropDown=False)
dv_type  = DataValidation(type="list",
    formula1='"Naturel,Labo"', showDropDown=False)
dv_niv   = DataValidation(type="list",
    formula1='"N1 Essentiel,N2 Sélection,N3 Prestige,N4 Excellence,N5 Exception,'
             'L1 Essentiel,L2 Sélection,L3 Prestige,L4 Excellence,L5 Exception"',
    showDropDown=False)
ws_cat.add_data_validation(dv_metal)
ws_cat.add_data_validation(dv_type)
ws_cat.add_data_validation(dv_niv)
dv_metal.sqref = "D4:D500"
dv_type.sqref  = "E4:E500"
dv_niv.sqref   = "F4:F500"

# ── Formules ──────────────────────────────────────────────────────────────────
def vlookup_group(mm_col, qt_col, r):
    """VLOOKUP dans TABLE_POIDS × quantité, 0 si cellule vide."""
    mm = f"{mm_col}{r}"
    qt = f"{qt_col}{r}"
    return (f'IF({mm}="",0,'
            f'IFERROR(VLOOKUP({mm},TABLE_POIDS!$A:$B,2,1),0)*IF({qt}="",0,{qt}))')

def cours_diam(type_col, niv_col, r):
    """Retourne le cours diamant selon type/niveau depuis COURS."""
    # COURS : B9=N1..B13=N5 (après header rows) / B17=L1..B21=L5
    # Section naturel commence ligne 5 (3 titre + 1 header + 1 data = ligne 5)
    # Recalcul selon la structure réelle de l'onglet COURS :
    # OR: ligne 4 (titre), 5 (header), 6 (data)
    # NATURELS: ligne 7 (titre), 8 (header), 9-13 (5 niveaux)
    # LABO: ligne 15 (titre), 16 (header), 17-21 (5 niveaux)
    tc = f"{type_col}{r}"
    nc = f"{niv_col}{r}"
    nat = (f'IF({nc}="N1 Essentiel",COURS!$B$9,'
           f'IF({nc}="N2 Sélection",COURS!$B$10,'
           f'IF({nc}="N3 Prestige",COURS!$B$11,'
           f'IF({nc}="N4 Excellence",COURS!$B$12,'
           f'IF({nc}="N5 Exception",COURS!$B$13,0)))))')
    lab = (f'IF({nc}="L1 Essentiel",COURS!$B$17,'
           f'IF({nc}="L2 Sélection",COURS!$B$18,'
           f'IF({nc}="L3 Prestige",COURS!$B$19,'
           f'IF({nc}="L4 Excellence",COURS!$B$20,'
           f'IF({nc}="L5 Exception",COURS!$B$21,0)))))')
    return f'IF({tc}="Naturel",{nat},{lab})'

def coeff_metal(metal_col, r):
    m = f"{metal_col}{r}"
    return (f'IF({m}="Or blanc",1+CONFIG!$B$6,'
            f'IF({m}="Or rose",1+CONFIG!$B$7,1))')

# ── Données exemples (issues de Classeur1) ───────────────────────────────────
# Format: (REF, NOM, COLLECTION, MÉTAL, TYPE, NIVEAU, POIDS_OR,
#          [(mm1,qty1),(mm2,qty2),...])
sample = [
    ("1001-OB-N", "Modèle 1001", "Pavé",      "Or blanc", "Naturel", "N3 Prestige", 3.8,
     [(5.20,1),(3.50,2),(1.80,6),(1.60,16),(1.50,24)]),
    ("1001-OB-L", "Modèle 1001", "Pavé",      "Or blanc", "Labo",    "L3 Prestige", 3.8,
     [(5.20,1),(3.50,2),(1.80,6),(1.60,16),(1.50,24)]),
    ("1002-OB-N", "Modèle 1002", "Pavé",      "Or blanc", "Naturel", "N2 Sélection", 4.1,
     [(6.50,1),(2.00,2),(1.10,48),(1.00,34),(0.80,26)]),
    ("1003-OB-N", "Modèle 1003", "Solitaires","Or blanc", "Naturel", "N3 Prestige", 3.2,
     [(2.20,2),(1.70,9),(1.40,20),(1.30,18),None]),
    ("1003-OJ-N", "Modèle 1003", "Solitaires","Or jaune", "Naturel", "N3 Prestige", 3.2,
     [(2.20,2),(1.70,9),(1.40,20),(1.30,18),None]),
    ("1004-OB-N", "Modèle 1004", "Alliances", "Or blanc", "Naturel", "N2 Sélection", 2.8,
     [(3.00,22),None,None,None,None]),
    ("1004-OR-L", "Modèle 1004", "Alliances", "Or rose",  "Labo",    "L2 Sélection", 2.8,
     [(3.00,22),None,None,None,None]),
]

G_COLS = [(7,8),(9,10),(11,12),(13,14),(15,16)]  # (mm_col, qty_col) index

for idx, (ref, nom, coll, metal, typ, niv, poids_or, groups) in enumerate(sample, 4):
    r = idx
    bg = OR_LIGHT if r % 2 == 0 else BLANC

    cel(ws_cat, r, COL["REF"],  ref,  bg=bg, bold=True)
    cel(ws_cat, r, COL["NOM"],  nom,  bg=bg)
    cel(ws_cat, r, COL["COLL"], coll, bg=bg)

    m_bg = {"Or blanc":BLEU_LIGHT,"Or jaune":OR_LIGHT,"Or rose":ROUGE_LIGHT}.get(metal,bg)
    cel(ws_cat, r, COL["METAL"], metal, bg=m_bg, bold=True, align="center")

    t_bg = BLEU_LIGHT if typ=="Naturel" else VIOLET_LIGHT
    cel(ws_cat, r, COL["TYPE"], typ, bg=t_bg, bold=True, align="center")
    cel(ws_cat, r, COL["NIV"],  niv, bg=bg, align="center")

    # Groupes diamants
    for gi, (mm_c, qt_c) in enumerate(G_COLS):
        g = groups[gi] if gi < len(groups) else None
        mm_val = g[0] if g else None
        qt_val = g[1] if g else None
        c = cel(ws_cat, r, mm_c, mm_val, bg=TEAL_LIGHT if mm_val else bg,
                align="center", fmt='0.00 "mm"' if mm_val else None)
        if mm_val: c.font = fnt(True, TEAL, 10)
        cel(ws_cat, r, qt_c, qt_val, bg=TEAL_LIGHT if qt_val else bg,
            align="center")

    # Carats calculés = somme VLOOKUP × qté sur les 5 groupes
    mm_lets = [get_column_letter(c[0]) for c in G_COLS]
    qt_lets = [get_column_letter(c[1]) for c in G_COLS]
    carats_f = "+".join(vlookup_group(mm_lets[i], qt_lets[i], r) for i in range(5))
    c = cel(ws_cat, r, COL["CARATS"], None, bg=VERT_LIGHT, bold=True, align="center",
            fmt='0.000 "ct"')
    c.value = f"=ROUND({carats_f},3)"

    c = cel(ws_cat, r, COL["POIDS_OR"], poids_or, bg=bg, align="center", fmt='0.0 "g"')

    # Main d'œuvre — pointe sur CONFIG par défaut, modifiable
    c = cel(ws_cat, r, COL["MO"], None, bg=OR_LIGHT, bold=True, align="center",
            fmt='#,##0 "€"')
    c.value = "=CONFIG!$B$9"

    # Coût or
    metal_l  = get_column_letter(COL["METAL"])
    poids_l  = get_column_letter(COL["POIDS_OR"])
    c = cel(ws_cat, r, COL["COUT_OR"], None, bg=ROUGE_LIGHT, align="center",
            fmt='#,##0.00 "€"')
    c.value = (f"=ROUND({poids_l}{r}*COURS!$B$6"
               f"*CONFIG!$B$5*{coeff_metal(metal_l,r)},2)")

    # Coût diamant
    type_l   = get_column_letter(COL["TYPE"])
    niv_l    = get_column_letter(COL["NIV"])
    carats_l = get_column_letter(COL["CARATS"])
    c = cel(ws_cat, r, COL["COUT_DIAM"], None, bg=ROUGE_LIGHT, align="center",
            fmt='#,##0.00 "€"')
    c.value = f"=ROUND({carats_l}{r}*{cours_diam(type_l,niv_l,r)},2)"

    # Prix de revient
    cout_or_l   = get_column_letter(COL["COUT_OR"])
    cout_diam_l = get_column_letter(COL["COUT_DIAM"])
    mo_l        = get_column_letter(COL["MO"])
    c = cel(ws_cat, r, COL["REVIENT"], None, bg=OR_LIGHT, bold=True, align="center",
            fmt='#,##0.00 "€"')
    c.value = f"=ROUND({cout_or_l}{r}+{cout_diam_l}{r}+{mo_l}{r}+CONFIG!$B$10,2)"

    # Prix HT
    rev_l = get_column_letter(COL["REVIENT"])
    c = cel(ws_cat, r, COL["HT"], None, bg=OR_LIGHT, bold=True, align="center",
            fmt='#,##0.00 "€"')
    c.value = f"=ROUND({rev_l}{r}*CONFIG!$B$12,2)"

    # Prix TTC
    ht_l = get_column_letter(COL["HT"])
    c = cel(ws_cat, r, COL["TTC"], None, bg=OR, bold=True, align="center",
            fmt='#,##0 "€"')
    c.value = f"=ROUND({ht_l}{r}*(1+CONFIG!$B$13),2)"

    ws_cat.row_dimensions[r].height = 20

# Ligne d'invitation
next_r = len(sample) + 4
ws_cat.merge_cells(f"A{next_r}:{get_column_letter(24)}{next_r}")
c = ws_cat[f"A{next_r}"]
c.value = ("⟶  Copiez une ligne existante pour ajouter un produit  —  "
           "Colonnes blanches = saisie manuelle  |  Colonnes vertes/rouges/dorées = calcul automatique")
c.fill = fill(GRIS); c.font = fnt(False,"888888",9,True); c.alignment = ctr()
ws_cat.row_dimensions[next_r].height = 18

# ═══════════════════════════════════════════════════════════════════════════════
# ONGLET 5 — EXPORT SHOPIFY
# ═══════════════════════════════════════════════════════════════════════════════
ws_exp = wb.create_sheet("EXPORT SHOPIFY")
ws_exp.sheet_properties.tabColor = VERT

exp_hdrs = [
    "Handle","Title","Body (HTML)","Vendor","Type","Tags","Published",
    "Option1 Name","Option1 Value","Option2 Name","Option2 Value",
    "Option3 Name","Option3 Value",
    "Variant SKU","Variant Price","Variant Compare At Price",
    "Variant Inventory Qty","Image Src","Image Alt Text",
    "SEO Title","SEO Description",
]
for i, w in enumerate([22,32,40]+[18]*(len(exp_hdrs)-3), 1):
    ws_exp.column_dimensions[get_column_letter(i)].width = w

ws_exp.merge_cells(f"A1:{get_column_letter(len(exp_hdrs))}1")
c = ws_exp["A1"]
c.value = "EXPORT SHOPIFY — Format Matrixify (CSV)"
c.fill = fill(VERT); c.font = fnt(True, BLANC, 13); c.alignment = ctr()
ws_exp.row_dimensions[1].height = 30

for i, h in enumerate(exp_hdrs, 1):
    hdr(ws_exp, 2, i, h, bg=NOIR, fg=BLANC, sz=9)
ws_exp.row_dimensions[2].height = 28
ws_exp.freeze_panes = "A3"

ws_exp.merge_cells(f"A3:{get_column_letter(len(exp_hdrs))}3")
c = ws_exp["A3"]
c.value = ("ℹ  Une ligne par taille de bague (même Handle).  "
           "Variant Price = reprenez le Prix TTC de l'onglet CATALOGUE.  "
           "Exportez en .CSV → importez via Matrixify dans Shopify.")
c.fill = fill(GRIS); c.font = fnt(False,NOIR,9,True); c.alignment = lft()
ws_exp.row_dimensions[3].height = 30

ex = ["1001-ob-n","Modèle 1001 — Or Blanc — Naturel","<p>Description SEO</p>",
      "Bijou-R","Bague Pavé","pavé, diamant naturel, or blanc",
      "TRUE","Taille","52","Métal","Or blanc","Diamant","N3 Prestige",
      "1001-OB-N-52","=CATALOGUE!X4","","3",
      "https://url-image.com/1001.jpg","Modèle 1001 pavé or blanc diamant naturel",
      "Bague Pavé Diamant Naturel Or Blanc — Bijou-R",
      "Découvrez notre bague pavé, composition unique de diamants naturels sertis en or blanc 18k."]
for i, v in enumerate(ex, 1):
    c = cel(ws_exp, 4, i, v, bg=GRIS)
    c.font = fnt(False,"555555",8,True)
ws_exp.row_dimensions[4].height = 18

# ── Sauvegarde ────────────────────────────────────────────────────────────────
out = "/home/user/Bijou-R/Catalogue_Bijou-R.xlsx"
wb.save(out)
print(f"OK — {out}")
